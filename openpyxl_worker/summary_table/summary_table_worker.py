import logging
from typing import List

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.constants import LEFT_TOP_ALIGN, THEME_RESULT_TABLE_HEADERS
from openpyxl_worker.summary_table.formatting import (
    apply_percentage_color_formatting,
    format_point_cells,
    set_borders,
)
from openpyxl_worker.types import (
    FormatArgs,
    MatrixCells,
    OverallResult,
    ResultCells,
    WorksheetRanges,
)


class SummaryTableWorker:
    """
    A worker class responsible for creating and formatting summary tables in Excel workbooks.
    This class handles the creation of summary tables with conditional formatting and proper styling.
    """

    # Column indices for the summary table
    NUMBER_COLUMN = 1
    TASK_NUMBER_COLUMN = 2
    THEME_COLUMN = 3
    POINT_COLUMN = 4

    def __init__(self, wb: Workbook, sheet_name: str) -> None:
        """
        Initialize the SummaryTableWorker.

        Args:
            wb: The workbook to work with
            sheet_name: Name of the worksheet to create or use
        """
        self.wb = wb
        self.ws = self._get_or_create_worksheet(sheet_name)

    def _get_or_create_worksheet(self, name: str) -> Worksheet:
        """Get existing worksheet or create a new one if it doesn't exist."""
        if name not in self.wb.sheetnames:
            logging.info("Creating new worksheet: %s", name)
            return self.wb.create_sheet(name)
        logging.info("Using existing worksheet: %s", name)
        return self.wb[name]

    def create(self, summary_table_data: List[WorksheetRanges]) -> None:
        """
        Create and format the summary table.

        Args:
            summary_table_data: List of worksheet ranges containing the data to summarize
        """
        self._add_header()
        result_cells = self._fill_table(summary_table_data)
        self._format_worksheet(result_cells)
        self._add_filter()

    def _add_header(self) -> None:
        """Add headers to the worksheet."""
        for index, header in enumerate(THEME_RESULT_TABLE_HEADERS, start=1):
            self.ws.cell(row=1, column=index, value=header)
        logging.info("Added summary table headers.")

    def _fill_table(self, summary_table_data: List[WorksheetRanges]) -> ResultCells:
        """
        Fill the table with data from the provided worksheet ranges.

        Args:
            summary_table_data: List of worksheet ranges containing the data

        Returns:
            ResultCells containing the filled table data
        """
        current_row = 2
        overall_results: List[OverallResult] = []

        for worksheet_data in summary_table_data:
            for index, task_cell in enumerate(worksheet_data.task_cells):
                row_data = self._create_row_data(
                    worksheet_data, task_cell, index, current_row
                )
                overall_results.append(row_data)
                current_row += 1

        logging.info("Filled summary table with %d rows.", len(overall_results))
        return ResultCells(self.ws.title, overall_results)

    def _create_row_data(
        self,
        worksheet_data: WorksheetRanges,
        task_cell: Cell,
        index: int,
        row: int,
    ) -> OverallResult:
        """Create a single row of data in the summary table."""
        number_cell = self.ws.cell(row=row, column=self.NUMBER_COLUMN, value=row - 1)

        task_number_cell = self.ws.cell(
            row=row,
            column=self.TASK_NUMBER_COLUMN,
            value=f"='{worksheet_data.name}'!{task_cell.coordinate}",
        )

        task_name_cell = self.ws.cell(
            row=row,
            column=self.THEME_COLUMN,
            value=f"='{worksheet_data.name}'!B{task_cell.row}",
        )

        percentage_cell = self.ws.cell(
            row=row,
            column=self.POINT_COLUMN,
            value=f"='{worksheet_data.name}'!{worksheet_data.percentage_of_completion_formulas[index].coordinate}",
        )

        return OverallResult(
            number_cell,
            task_number_cell,
            task_name_cell,
            percentage_cell,
        )

    def _format_worksheet(self, result_cells: ResultCells) -> None:
        """
        Apply formatting to the worksheet including conditional formatting and borders.

        Args:
            result_cells: The cells to format
        """
        # Remove all conditional formatting rules (openpyxl does not have a clear() method)
        self.ws.conditional_formatting._cf_rules.clear()  # type: ignore[attr-defined]

        # Format task number cells
        task_number_cells = tuple(
            cells.task_number for cells in result_cells.overall_result
        )
        # Use a default FormatArgs for left-top alignment
        format_point_cells(
            task_number_cells, format_args=FormatArgs(alignment=LEFT_TOP_ALIGN)
        )

        # Set borders for the entire table
        table_cells = self._get_table_cells(result_cells)
        set_borders(table_cells)

        # Apply percentage color formatting
        if result_cells.overall_result:
            start_coordinate = result_cells.overall_result[
                0
            ].percentage_of_completion.coordinate
            end_coordinate = result_cells.overall_result[
                -1
            ].percentage_of_completion.coordinate
            apply_percentage_color_formatting(self.ws, start_coordinate, end_coordinate)

        logging.info("Formatted summary worksheet: %s", self.ws.title)

    def _get_table_cells(self, result_cells: ResultCells) -> MatrixCells:
        """Get all cells that should be included in the table formatting."""
        # if not result_cells.overall_result:
        #     return tuple()
        first_row = result_cells.overall_result[0]
        last_row = result_cells.overall_result[-1]
        return tuple(
            self.ws.iter_rows(
                min_row=first_row.number.row,
                max_row=last_row.number.row,
                min_col=self.NUMBER_COLUMN,
                max_col=self.POINT_COLUMN,
            )
        )

    def _add_filter(self) -> None:
        """Add auto-filter to the entire worksheet."""
        self.ws.auto_filter.ref = self.ws.dimensions
        logging.info("Added auto-filter to worksheet: %s", self.ws.title)
