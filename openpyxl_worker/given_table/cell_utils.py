import logging
from typing import List, Tuple

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.given_table.constants import EMPTY_STUDENT, REPLACE_VALUES
from openpyxl_worker.types import (
    FilledRows,
    FinderCells,
    LineCells,
    MatrixCells,
    Range,
)


def get_nonempty_rows(ws: Worksheet, cell_range: Range) -> FilledRows:
    """Return rows in the worksheet that are not empty according to EMPTY_STUDENT.

    Args:
        ws (Worksheet): The worksheet to process.
        cell_range (Range): The cell range to consider.

    Returns:
        FilledRows: Tuple of filled rows and the last row number.
    """
    start_cell: Cell = ws[cell_range.start]
    end_cell: Cell = ws[cell_range.end]
    rows: List[Tuple[Cell, ...]] = []
    last_row_number = end_cell.row

    for row in ws.iter_rows(
        min_row=start_cell.row,
        max_row=end_cell.row,
        min_col=start_cell.column,
        max_col=end_cell.column,
    ):
        cell = ws.cell(row[0].row, row[0].column - 1)
        if cell.value != EMPTY_STUDENT:
            rows.append(row)

    logging.info(
        "Selected %d filled rows in range %s:%s",
        len(rows),
        cell_range.start,
        cell_range.end,
    )
    return FilledRows(tuple(rows), last_row_number)


def replace_cells_with_zero(cell_range: MatrixCells) -> MatrixCells:
    """Replace cells with values in REPLACE_VALUES with 0 in the given matrix.

    Args:
        cell_range (MatrixCells): The matrix of cells to process.

    Returns:
        MatrixCells: The updated matrix with replacements.
    """
    for row in cell_range:
        for cell in row:
            if cell.value in REPLACE_VALUES:
                cell.value = 0
    logging.info("Replaced X cells in matrix.")
    return cell_range


def remove_variant_columns(
    ws: Worksheet, cell_range: MatrixCells, variant_str: str
) -> MatrixCells:
    """Remove columns where the header contains the variant string.

    Args:
        ws (Worksheet): The worksheet to process.
        cell_range (MatrixCells): The matrix of cells to process.
        variant_str (str): The string to identify variant columns.

    Returns:
        MatrixCells: The updated matrix with variant columns removed.
    """
    matrix_cells: List[Tuple[Cell, ...]] = []
    for row in cell_range:
        row_cells: List[Cell] = []
        for cell in row:
            cell_column = cell.column
            header_cell = ws.cell(row=1, column=cell_column)
            if (
                isinstance(header_cell.value, str)
                and variant_str in header_cell.value.lower()
            ):
                continue
            row_cells.append(cell)
        matrix_cells.append(tuple(row_cells))
    logging.info("Removed variant columns in matrix.")
    return tuple(matrix_cells)


def extract_student_cells(ws: Worksheet, point_cells: MatrixCells) -> LineCells:
    """Extract student cells for each row in the matrix.

    Args:
        ws (Worksheet): The worksheet to process.
        point_cells (MatrixCells): The matrix of point cells.

    Returns:
        LineCells: Tuple of student cells.
    """
    student_cells = [ws.cell(row=row[0].row, column=1) for row in point_cells]
    return tuple(student_cells)


def extract_task_cells(ws: Worksheet, point_cells: MatrixCells) -> LineCells:
    """Extract task cells for the first row in the matrix.

    Args:
        ws (Worksheet): The worksheet to process.
        point_cells (MatrixCells): The matrix of point cells.

    Returns:
        LineCells: Tuple of task cells.
    """
    tasks_row = 1
    start_column = point_cells[0][0].column
    end_column = point_cells[0][-1].column
    return tuple(ws.cell(tasks_row, col) for col in range(start_column, end_column + 1))


def extract_student_and_task_cells(
    ws: Worksheet, cell_range: MatrixCells
) -> FinderCells:
    """Extract student and task cells for the given matrix.

    Args:
        ws (Worksheet): The worksheet to process.
        cell_range (MatrixCells): The matrix of cells.

    Returns:
        FinderCells: Named tuple of student and task cells.
    """
    student_cells = extract_student_cells(ws, cell_range)
    task_cells = extract_task_cells(ws, cell_range)
    return FinderCells(student_cells, task_cells)
