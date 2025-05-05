import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet


class WorkbookContainer:
    """Container for managing an Excel workbook and its worksheets.

    Provides methods to load, activate, and save workbooks and worksheets for processing.
    """

    path: Path
    wb: Workbook
    ws: OpenpyxlWorksheet

    def __init__(self, file_path: Path) -> None:
        """Initialize the workbook container by loading the workbook.

        Args:
            file_path (Path): Path to the Excel workbook file.
        Raises:
            Exception: If the workbook cannot be loaded.
        """
        self.path = file_path
        try:
            self.wb = load_workbook(self.path)
            self.ws = self.wb[self.wb.sheetnames[0]]
            logging.info("Loaded workbook: %s", self.path)
        except Exception:
            logging.exception("Failed to load workbook: %s", self.path)
            raise

    def activate_sheet(self, sheet_name: str) -> "WorkbookContainer":
        """Activate a worksheet by name.

        Args:
            sheet_name (str): Name of the worksheet to activate.

        Returns:
            WorkbookContainer: Self, with the active worksheet set.
        Raises:
            KeyError: If the worksheet name does not exist in the workbook.
        """
        try:
            self.ws = self.wb[sheet_name]
            logging.info("Activated worksheet: %s in %s", sheet_name, self.path)
        except KeyError:
            logging.error(
                "Worksheet '%s' not found in workbook '%s'", sheet_name, self.path
            )
            raise
        return self

    def save_table(self, file_path: Path) -> None:
        """Save the workbook to the specified file path.

        Args:
            file_path (Path): Path to save the workbook.
        Raises:
            Exception: If the workbook cannot be saved.
        """
        try:
            self.wb.save(file_path)
            logging.info("Workbook saved to: %s", file_path)
        except Exception:
            logging.exception("Failed to save workbook to: %s", file_path)
            raise
