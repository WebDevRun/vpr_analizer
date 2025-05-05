from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule
from openpyxl.styles import Alignment, Color
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.constants import (
    BRICK_COLOR,
    LEFT_TOP_ALIGN,
    LIME_COLOR,
    RIGHT_TOP_ALIGN,
    THEME_TABLE_HEADERS,
    THIN_BORDER,
    YELLOW_COLOR,
)
from openpyxl_worker.types import (
    FormatArgs,
    GivenTableCells,
    LineCells,
    MatrixCells,
    NumberFormatCell,
    WorksheetRanges,
)


class AnalyticTableCreates:
    """Class for creating and formatting analytic tables in Excel workbooks."""

    def __init__(self, wb: Workbook, ws: Worksheet, ranges: GivenTableCells) -> None:
        """Initialize AnalyticTableCreates.

        Args:
            wb (Workbook): The workbook to work with.
            ws (Worksheet): The worksheet to work with.
            ranges (GivenTableCells): The cell ranges and values for the table.
        """
        self.wb: Workbook = wb
        self.ws: Worksheet = ws
        self.ranges: GivenTableCells = ranges

    def create(self) -> WorksheetRanges:
        """Create and format the analytic table, returning worksheet ranges."""
        worksheet_ranges = self.create_table()
        self.format_worksheet(worksheet_ranges)
        self.paint_worksheet(worksheet_ranges)
        return worksheet_ranges

    def create_table(self) -> WorksheetRanges:
        """Create the table and return worksheet ranges. (Implementation omitted for brevity)"""
        table_headers = self.fill_table_header(
            self.ranges.student_cells, self.ranges.last_row
        )
        task_cells = self.fill_task_numbers(self.ranges.task_numbers, table_headers[0])
        task_description_cells = self.find_task_descriptions(
            task_cells, table_headers[1]
        )
        max_point_cells = self.fill_max_points(self.ranges.max_points, table_headers[2])
        student_formulas = self.find_student_formulas(table_headers)
        point_formulas = self.fill_point_formulas(
            self.ranges.point_cells, student_formulas
        )
        average_formulas = self.fill_average_formulas(point_formulas, table_headers[-2])
        percentage_of_completion_formulas = self.fill_percentage_of_completion(
            average_formulas,
            table_headers[2],
            table_headers[-1],
        )
        sum_max_point_formula = self.fill_sum_max_points(max_point_cells)
        sum_student_point_formulas = self.fill_sum_student_points(point_formulas)
        average_point = self.fill_average_point(average_formulas)
        average_percentage_of_completion = self.fill_average_percentage_of_completion(
            sum_student_point_formulas,
            sum_max_point_formula,
            percentage_of_completion_formulas[-1],
        )
        percentage_of_points = self.fill_percentage_of_point(
            sum_student_point_formulas, sum_max_point_formula
        )
        return WorksheetRanges(
            self.ws.title,
            table_headers,
            task_cells,
            point_formulas,
            average_formulas,
            percentage_of_completion_formulas,
            max_point_cells,
            average_point,
            average_percentage_of_completion,
            percentage_of_points,
            task_description_cells,
        )

    def fill_table_header(self, student_cells: LineCells, last_row: int) -> LineCells:
        start_row = last_row + 2
        start_column = 1
        headers = (
            THEME_TABLE_HEADERS.number,
            THEME_TABLE_HEADERS[1],
            THEME_TABLE_HEADERS[2],
            *[f"={cell.coordinate}" for cell in student_cells],
            *THEME_TABLE_HEADERS[3:],
        )
        filled_cells = [
            self.ws.cell(start_row, start_column + index, header)
            for index, header in enumerate(headers)
        ]
        return tuple(filled_cells)

    def fill_task_numbers(
        self, task_cells: Tuple[str, ...], number_cell: Cell
    ) -> LineCells:
        start_cell = self.ws.cell(number_cell.row + 1, number_cell.column)
        filled_cells = [
            self.ws.cell(start_cell.row + index, start_cell.column, value)
            for index, value in enumerate(task_cells)
        ]
        return tuple(filled_cells)

    def find_task_descriptions(
        self, task_cells: LineCells, task_description_cell: Cell
    ) -> LineCells:
        start_cell = self.ws.cell(
            task_description_cell.row + 1, task_description_cell.column
        )
        found_cells = [
            self.ws.cell(start_cell.row + index, start_cell.column)
            for index, _ in enumerate(task_cells)
        ]
        return tuple(found_cells)

    def fill_max_points(
        self, max_point_cells: Tuple[int, ...], max_point_cell: Cell
    ) -> LineCells:
        start_cell = self.ws.cell(max_point_cell.row + 1, max_point_cell.column)
        filled_cells = [
            self.ws.cell(start_cell.row + index, start_cell.column, value)
            for index, value in enumerate(max_point_cells)
        ]
        return tuple(filled_cells)

    def find_student_formulas(self, header_cells: LineCells) -> LineCells:
        return tuple(
            cell for cell in header_cells if cell.value not in THEME_TABLE_HEADERS
        )

    def fill_point_formulas(
        self, point_cells: MatrixCells, student_cells: LineCells
    ) -> MatrixCells:
        filled_cells: List[Tuple[Cell, ...]] = []

        for index, point_cell_row in enumerate(point_cells):
            column = student_cells[index].column
            start_row = student_cells[index].row + 1
            filled_cells.append(
                tuple(
                    self.ws.cell(start_row + j, column, f"={point_cell.coordinate}")
                    for j, point_cell in enumerate(point_cell_row)
                )
            )

        return tuple(filled_cells)

    def fill_average_formulas(
        self, point_cells: MatrixCells, average_cell: Cell
    ) -> LineCells:
        column = average_cell.column
        start_row = average_cell.row + 1
        filled_cells: List[Cell] = []

        for index, start_cell in enumerate(point_cells[0]):
            end_cell = point_cells[-1][index]
            cell_formula = f"=AVERAGE({start_cell.coordinate}:{end_cell.coordinate})"
            cell = self.ws.cell(start_row + index, column, cell_formula)
            filled_cells.append(cell)

        return tuple(filled_cells)

    def fill_percentage_of_completion(
        self,
        average_cells: LineCells,
        max_point_cell: Cell,
        percentage_of_completion_cell: Cell,
    ) -> LineCells:
        column = percentage_of_completion_cell.column
        start_row = percentage_of_completion_cell.row + 1
        filled_cells: List[Cell] = []

        for index, cell in enumerate(average_cells):
            max_point_coordinate = f"{max_point_cell.column_letter}{cell.row}"
            cell_formula = f"={cell.coordinate}/{max_point_coordinate}"
            cell = self.ws.cell(start_row + index, column, cell_formula)
            filled_cells.append(cell)

        return tuple(filled_cells)

    def fill_sum_max_points(self, max_point_cells: LineCells):
        last_cell = max_point_cells[-1]
        row = last_cell.row + 1
        column = last_cell.column
        sum_max_point_formula = self.ws.cell(row, column)
        sum_max_point_formula.value = (
            f"=SUM({max_point_cells[0].coordinate}:{max_point_cells[-1].coordinate}"
        )
        return sum_max_point_formula

    def fill_sum_student_points(self, student_cells: MatrixCells):
        filled_cells: List[Cell] = []

        for student_cell_row in student_cells:
            last_cell = student_cell_row[-1]
            row = last_cell.row + 1
            column = last_cell.column
            sum_student_point_formula = self.ws.cell(row, column)
            sum_student_point_formula.value = f"=SUM({student_cell_row[0].coordinate}:{student_cell_row[-1].coordinate})"
            filled_cells.append(sum_student_point_formula)

        return tuple(filled_cells)

    def fill_average_point(self, average_formulas: Tuple[Cell, ...]):
        last_cell = average_formulas[-1]
        row = last_cell.row + 1
        column = last_cell.column
        cell = self.ws.cell(row, column)
        cell.value = f"=AVERAGE({average_formulas[0].coordinate}:{average_formulas[-1].coordinate})"
        return cell

    def fill_average_percentage_of_completion(
        self,
        sum_student_point: Tuple[Cell, ...],
        sum_max_point: Cell,
        last_percentage_of_completion: Cell,
    ):
        row = last_percentage_of_completion.row + 1
        column = last_percentage_of_completion.column
        cell = self.ws.cell(row, column)
        average_chunk = f"=AVERAGE({sum_student_point[0].coordinate}:{sum_student_point[-1].coordinate})"
        cell.value = f"{average_chunk}/{sum_max_point.coordinate}"
        return cell

    def fill_percentage_of_point(
        self,
        sum_student_point: Tuple[Cell, ...],
        sum_max_point: Cell,
    ):
        filled_cells = (
            self.ws.cell(
                cell.row + 1,
                cell.column,
                f"={cell.coordinate}/{sum_max_point.coordinate}",
            )
            for cell in sum_student_point
        )
        return tuple(filled_cells)

    def format_worksheet(self, worksheet_ranges: WorksheetRanges) -> None:
        """Format the worksheet with borders, alignment, and number formats."""
        table_header_format = FormatArgs(LEFT_TOP_ALIGN, wrap_text=True)
        number_formula_format = FormatArgs(LEFT_TOP_ALIGN)
        point_formula_format = FormatArgs(RIGHT_TOP_ALIGN)
        average_formula_format = FormatArgs(
            RIGHT_TOP_ALIGN,
            NumberFormatCell.FORMAT_NUMBER_00,
        )
        percentage_formula_format = FormatArgs(
            RIGHT_TOP_ALIGN, NumberFormatCell.FORMAT_PERCENTAGE_00
        )
        self.format_not_point_cells(worksheet_ranges.table_headers, table_header_format)
        self.format_not_point_cells(worksheet_ranges.task_cells, number_formula_format)
        self.format_point_cells(worksheet_ranges.point_formulas, point_formula_format)
        self.format_not_point_cells(
            worksheet_ranges.average_formulas, average_formula_format
        )
        self.format_not_point_cells(
            worksheet_ranges.percentage_of_completion_formulas,
            percentage_formula_format,
        )
        self.format_not_point_cells(
            (worksheet_ranges.average_point,), average_formula_format
        )
        self.format_not_point_cells(
            (worksheet_ranges.average_percentage_of_completion,),
            percentage_formula_format,
        )
        self.format_not_point_cells(
            worksheet_ranges.percentage_of_points, percentage_formula_format
        )
        table_cells = self.find_table_cells(
            worksheet_ranges.table_headers, worksheet_ranges.percentage_of_points
        )
        self.set_borders(table_cells)

    def format_not_point_cells(
        self, cells: Tuple[Cell, ...], format_args: FormatArgs
    ) -> Tuple[Cell, ...]:
        """Format non-point cells with alignment and number format."""
        for cell in cells:
            cell.alignment = Alignment(
                format_args.alignment.horizontal,
                format_args.alignment.vertical,
                wrap_text=format_args.wrap_text,
            )
            if format_args.number_format.value:
                cell.number_format = format_args.number_format.value
        return cells

    def format_point_cells(
        self, cells: MatrixCells, format_args: FormatArgs
    ) -> MatrixCells:
        """Format point cells in a matrix with alignment and number format."""
        for row in cells:
            self.format_not_point_cells(row, format_args)
        return cells

    def set_borders(self, cells: MatrixCells) -> MatrixCells:
        """Set thin borders for all cells in the matrix."""
        for row in cells:
            for cell in row:
                cell.border = THIN_BORDER
        return cells

    def generate_percentage_color_rule(self) -> Rule:
        """Generate a color scale rule for percentage formatting."""
        first = FormatObject(type="num", val=0)
        mid = FormatObject(type="num", val=0.5)
        last = FormatObject(type="num", val=1)
        colors = [
            Color(BRICK_COLOR),
            Color(YELLOW_COLOR),
            Color(LIME_COLOR),
        ]
        color_scale = ColorScale(cfvo=[first, mid, last], color=colors)
        return Rule(type="colorScale", colorScale=color_scale)

    def generate_point_color_rule(self, max_point: int) -> Rule:
        """Generate a color scale rule for point formatting based on max point."""
        first = FormatObject(type="num", val=0)
        mid = FormatObject(type="num", val=max_point / 2)
        last = FormatObject(type="num", val=max_point)
        colors = [
            Color(BRICK_COLOR),
            Color(YELLOW_COLOR),
            Color(LIME_COLOR),
        ]
        color_scale = ColorScale(cfvo=[first, mid, last], color=colors)
        return Rule(type="colorScale", colorScale=color_scale)

    def find_table_cells(
        self,
        horizontal_cells: LineCells,
        vertical_cells: LineCells,
    ):
        start_column = horizontal_cells[0].column
        end_column = horizontal_cells[-1].column
        start_row = horizontal_cells[0].row
        end_row = vertical_cells[-1].row

        return tuple(
            self.ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_column,
                max_col=end_column,
            )
        )

    def paint_worksheet(self, worksheet_ranges: WorksheetRanges) -> None:
        """Apply color formatting to the worksheet."""
        self.ws.conditional_formatting = ConditionalFormattingList()
        percent_color_rule = self.generate_percentage_color_rule()
        start_coordinate = worksheet_ranges.percentage_of_points[0].coordinate
        end_coordinate = worksheet_ranges.percentage_of_points[-1].coordinate
        self.ws.conditional_formatting.add(
            f"{start_coordinate}:{end_coordinate}",
            percent_color_rule,
        )
        start_coordinate = worksheet_ranges.percentage_of_completion_formulas[
            0
        ].coordinate
        end_coordinate = worksheet_ranges.average_percentage_of_completion.coordinate
        self.ws.conditional_formatting.add(
            f"{start_coordinate}:{end_coordinate}",
            percent_color_rule,
        )

        for index, cell in enumerate(worksheet_ranges.max_point_cells):
            if type(cell.value) is not int:
                raise ValueError("Cell value is not integer")

            point_color_rule = self.generate_point_color_rule(cell.value)
            row = cell.row
            start_column = cell.column + 1
            end_column = worksheet_ranges.average_formulas[index].column - 1
            start_cell = self.ws.cell(row, start_column)
            end_cell = self.ws.cell(row, end_column)
            self.ws.conditional_formatting.add(
                f"{start_cell.coordinate}:{end_cell.coordinate}",
                point_color_rule,
            )
