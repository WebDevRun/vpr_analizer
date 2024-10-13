from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.constants import THEME_TABLE_HEADERS
from openpyxl_worker.types import LineCells, MatrixCells, Range


class CellsFinder:
    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws

    def getCells(self, cell_range: Range) -> MatrixCells:
        start_cell = self.ws[cell_range.start]
        end_cell = self.ws[cell_range.end]

        return tuple(
            tuple(
                self.ws.cell(row, col)
                for col in range(start_cell.column, end_cell.column + 1)
            )
            for row in range(start_cell.row, end_cell.row + 1)
        )

    def find_students(self, point_cells: MatrixCells) -> LineCells:
        start_row = point_cells[0][0].row
        end_row = point_cells[-1][0].row
        student_cells = [self.ws.cell(row, 1) for row in range(start_row, end_row + 1)]
        return tuple(student_cells)

    def find_task_cells(self, point_cells: MatrixCells) -> LineCells:
        tasks_row = point_cells[0][0].row - 1
        start_column = point_cells[0][0].column
        end_column = point_cells[0][-1].column
        return tuple(
            self.ws.cell(tasks_row, col) for col in range(start_column, end_column + 1)
        )

    def find_student_formulas(self, header_cells: LineCells) -> LineCells:
        return tuple(
            cell for cell in header_cells if cell.value not in THEME_TABLE_HEADERS
        )

    def find_table_cells(
        self,
        horizontal_cells: LineCells,
        vertical_cells: LineCells,
    ):
        start_column = horizontal_cells[0].column
        end_column = horizontal_cells[-1].column
        start_row = horizontal_cells[0].row
        end_row = vertical_cells[-1].row

        return tuple(
            self.ws.iter_rows(
                min_row=start_row,
                max_row=end_row,
                min_col=start_column,
                max_col=end_column,
            )
        )

    def find_max_point_cells(
        self, number_formulas: LineCells, table_header: Cell
    ) -> LineCells:
        row = table_header.row + 1
        column = table_header.column
        start_cell = self.ws.cell(row, column)
        end_cell = self.ws.cell(number_formulas[-1].row, column)
        cells = (
            self.ws.cell(row, column) for row in range(start_cell.row, end_cell.row + 1)
        )

        return tuple(cells)
