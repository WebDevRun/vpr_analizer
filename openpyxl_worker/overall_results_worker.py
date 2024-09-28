from dataclasses import astuple
from typing import List

from openpyxl import Workbook
from openpyxl.formatting.formatting import ConditionalFormattingList

from openpyxl_worker.cell_formater import CellFormater
from openpyxl_worker.cells_finder import CellsFinder
from openpyxl_worker.constants import THEME_RESULT_TABLE_HEADERS
from openpyxl_worker.types import (
    FormatArgs,
    OverallResult,
    ResultCells,
    WorksheetRanges,
)


class OverallResultsWorker:
    NUMBER_COLUMN = 1
    CLASS_NUMBER = 2
    TASK_NUMBER_COLUMN = 3
    THEME_COLUMN = 4
    POINT_COLUMN = 5

    def __init__(self, wb: Workbook, name: str):
        self.wb = wb

        if name not in self.wb.sheetnames:
            self.ws = self.wb.create_sheet(name)
            return

        self.ws = self.wb[name]

    def fill_table(self, wb_ranges: List[WorksheetRanges]):
        self.add_header()
        result_cells = self.calculate_values(wb_ranges)
        self.format_ws(result_cells)

    def add_header(self):
        start_row = 1

        for str in THEME_RESULT_TABLE_HEADERS:
            column = THEME_RESULT_TABLE_HEADERS.index(str) + 1
            self.ws.cell(row=start_row, column=column, value=str)

    def calculate_values(self, wb_ranges: List[WorksheetRanges]):
        start_row = 2
        ws_list: List[OverallResult] = []

        for wb in wb_ranges:
            for index, cell in enumerate(wb.task_formulas):
                number = self.ws.cell(
                    row=start_row, column=self.NUMBER_COLUMN, value=start_row - 1
                )
                class_number = self.ws.cell(
                    row=start_row,
                    column=self.CLASS_NUMBER,
                    value=wb.name[0],
                )
                cell_value = f"='{wb.name}'!{cell.coordinate}"
                task_number = self.ws.cell(
                    row=start_row,
                    column=self.TASK_NUMBER_COLUMN,
                    value=cell_value,
                )
                cell_value = f"='{wb.name}'!B{cell.row}"
                task_name = self.ws.cell(
                    row=start_row,
                    column=self.THEME_COLUMN,
                    value=cell_value,
                )
                cell_value = f"='{wb.name}'!{wb.percentage_of_completion_formulas[index].coordinate}"
                percentage_of_completion = self.ws.cell(
                    row=start_row,
                    column=self.POINT_COLUMN,
                    value=cell_value,
                )
                start_row += 1
                overall_result = OverallResult(
                    number,
                    class_number,
                    task_number,
                    task_name,
                    percentage_of_completion,
                )
                ws_list.append(overall_result)

        return ResultCells(self.ws.title, ws_list)

    def format_ws(self, result_cells: ResultCells):
        cell_formater = CellFormater()
        cell_finder = CellsFinder(self.ws)
        self.ws.conditional_formatting = ConditionalFormattingList()
        number_task_format = FormatArgs(cell_formater.LEFT_TOP_ALIGN)
        task_number_cells = tuple(
            [cells.task_number for cells in result_cells.overall_result]
        )
        cells = cell_formater.format_not_point_cells(
            task_number_cells, number_task_format
        )
        horizontal_cells = astuple(result_cells.overall_result[0])
        table_cells = cell_finder.find_table_cells(horizontal_cells, task_number_cells)
        cell_formater.set_borders(table_cells)
        percent_color_rule = cell_formater.generate_percentage_color_rule()
        start_coordinate = result_cells.overall_result[
            0
        ].percentage_of_completion.coordinate
        end_coordinate = result_cells.overall_result[
            -1
        ].percentage_of_completion.coordinate
        self.ws.conditional_formatting.add(
            f"{start_coordinate}:{end_coordinate}",
            percent_color_rule,
        )
        return cells
