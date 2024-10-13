from typing import List, Tuple

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.cells_finder import MatrixCells
from openpyxl_worker.constants import THEME_TABLE_HEADERS
from openpyxl_worker.types import LineCells


class CellsFiller:
    def __init__(self, ws: Worksheet):
        self.ws = ws

    def fill_table_header(self, student_cells: LineCells) -> LineCells:
        start_row = student_cells[-1].row + 2
        start_column = 1
        headers = (
            THEME_TABLE_HEADERS.number,
            THEME_TABLE_HEADERS[1],
            THEME_TABLE_HEADERS[2],
            *[f"={cell.coordinate}" for cell in student_cells],
            *THEME_TABLE_HEADERS[3:],
        )
        filled_cells = [
            self.ws.cell(start_row, start_column + index, header)
            for index, header in enumerate(headers)
        ]
        return tuple(filled_cells)

    def fill_task_numbers(self, task_cells: LineCells, number_cell: Cell) -> LineCells:
        start_cell = self.ws.cell(number_cell.row + 1, number_cell.column)
        filled_cells = [
            self.ws.cell(
                start_cell.row + index, start_cell.column, f"={cell.coordinate}"
            )
            for index, cell in enumerate(task_cells)
        ]
        return tuple(filled_cells)

    def fill_point_formulas(
        self, point_cells: MatrixCells, student_cells: LineCells
    ) -> MatrixCells:
        filled_cells: List[Tuple[Cell, ...]] = []

        for index, point_cell_row in enumerate(point_cells):
            column = student_cells[index].column
            start_row = student_cells[index].row + 1
            filled_cells.append(
                tuple(
                    self.ws.cell(start_row + j, column, f"={point_cell.coordinate}")
                    for j, point_cell in enumerate(point_cell_row)
                )
            )

        return tuple(filled_cells)

    def fill_average_formulas(
        self, point_cells: MatrixCells, average_cell: Cell
    ) -> LineCells:
        column = average_cell.column
        start_row = average_cell.row + 1
        filled_cells: List[Cell] = []

        for index, start_cell in enumerate(point_cells[0]):
            end_cell = point_cells[-1][index]
            cell_formula = f"=AVERAGE({start_cell.coordinate}:{end_cell.coordinate})"
            cell = self.ws.cell(start_row + index, column, cell_formula)
            filled_cells.append(cell)

        return tuple(filled_cells)

    def fill_percentage_of_completion(
        self,
        average_cells: LineCells,
        max_point_cell: Cell,
        percentage_of_completion_cell: Cell,
    ) -> LineCells:
        column = percentage_of_completion_cell.column
        start_row = percentage_of_completion_cell.row + 1
        filled_cells: List[Cell] = []

        for index, cell in enumerate(average_cells):
            max_point_coordinate = f"{max_point_cell.column_letter}{cell.row}"
            cell_formula = f"={cell.coordinate}/{max_point_coordinate}"
            cell = self.ws.cell(start_row + index, column, cell_formula)
            filled_cells.append(cell)

        return tuple(filled_cells)

    def fill_sum_max_points(self, max_point_cells: LineCells):
        last_cell = max_point_cells[-1]
        row = last_cell.row + 1
        column = last_cell.column
        sum_max_point_formula = self.ws.cell(row, column)
        sum_max_point_formula.value = (
            f"=SUM({max_point_cells[0].coordinate}:{max_point_cells[-1].coordinate}"
        )
        return sum_max_point_formula

    def fill_sum_student_points(self, student_cells: MatrixCells):
        filled_cells: List[Cell] = []

        for student_cell_row in student_cells:
            last_cell = student_cell_row[-1]
            row = last_cell.row + 1
            column = last_cell.column
            sum_student_point_formula = self.ws.cell(row, column)
            sum_student_point_formula.value = f"=SUM({student_cell_row[0].coordinate}:{student_cell_row[-1].coordinate})"
            filled_cells.append(sum_student_point_formula)

        return tuple(filled_cells)

    def fill_average_point(self, average_formulas: Tuple[Cell, ...]):
        last_cell = average_formulas[-1]
        row = last_cell.row + 1
        column = last_cell.column
        cell = self.ws.cell(row, column)
        cell.value = f"=AVERAGE({average_formulas[0].coordinate}:{average_formulas[-1].coordinate})"
        return cell

    def fill_average_percentage_of_completion(
        self,
        sum_student_point: Tuple[Cell, ...],
        sum_max_point: Cell,
        last_percentage_of_completion: Cell,
    ):
        row = last_percentage_of_completion.row + 1
        column = last_percentage_of_completion.column
        cell = self.ws.cell(row, column)
        average_chunk = f"=AVERAGE({sum_student_point[0].coordinate}:{sum_student_point[-1].coordinate})"
        cell.value = f"{average_chunk}/{sum_max_point.coordinate}"
        return cell

    def fill_percentage_of_point(
        self, sum_student_point: Tuple[Cell, ...], sum_max_point: Cell
    ):
        filled_cells = (
            self.ws.cell(
                cell.row + 1,
                cell.column,
                f"={cell.coordinate}/{sum_max_point.coordinate}",
            )
            for cell in sum_student_point
        )
        return tuple(filled_cells)
