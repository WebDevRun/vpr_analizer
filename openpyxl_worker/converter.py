from typing import Dict

from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl_worker.cells_finder import CellsFinder
from openpyxl_worker.types import LineCells, MatrixCells, Range, WorksheetRanges
from yaml_worker.types import WorksheetStrRanges


class RangeConverter:
    SPLIT_SIMBOL = ":"

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws

    def ranges_to_str(self, worksheet_ranges: WorksheetRanges) -> WorksheetStrRanges:
        table_headers = self.get_str_from_line_cells(worksheet_ranges.table_headers)
        task_formulas = self.get_str_from_line_cells(worksheet_ranges.task_formulas)
        student_formulas = self.get_str_from_line_cells(
            worksheet_ranges.student_formulas
        )
        point_formulas = self.get_str_from_matrix_cells(worksheet_ranges.point_formulas)
        average_formulas = self.get_str_from_line_cells(
            worksheet_ranges.average_formulas
        )
        percentage_of_completion_formulas = self.get_str_from_line_cells(
            worksheet_ranges.percentage_of_completion_formulas
        )
        max_point_cells = self.get_str_from_line_cells(worksheet_ranges.max_point_cells)
        sum_student_point_formulas = self.get_str_from_line_cells(
            worksheet_ranges.sum_student_point_formulas
        )
        percentage_of_points = self.get_str_from_line_cells(
            worksheet_ranges.percentage_of_points
        )

        return WorksheetStrRanges(
            worksheet_ranges.name,
            table_headers,
            task_formulas,
            student_formulas,
            point_formulas,
            average_formulas,
            percentage_of_completion_formulas,
            max_point_cells,
            worksheet_ranges.sum_max_point_formula.coordinate,
            sum_student_point_formulas,
            worksheet_ranges.average_point.coordinate,
            worksheet_ranges.average_percentage_of_completion.coordinate,
            percentage_of_points,
        )

    def get_str_from_line_cells(self, line_cells: LineCells) -> str:
        return f"{line_cells[0].coordinate}:{line_cells[-1].coordinate}"

    def get_str_from_matrix_cells(self, matrix_cells: MatrixCells) -> str:
        return f"{matrix_cells[0][0].coordinate}:{matrix_cells[-1][-1].coordinate}"

    def str_to_ranges(self, worksheet_str_ranges: Dict[str, str]) -> WorksheetRanges:
        cell_finder = CellsFinder(self.ws)
        coordinate = worksheet_str_ranges["table_headers"].split(self.SPLIT_SIMBOL)
        table_headers = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        coordinate = worksheet_str_ranges["task_formulas"].split(self.SPLIT_SIMBOL)
        task_formulas = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        task_formulas = tuple([row[0] for row in task_formulas])
        coordinate = worksheet_str_ranges["student_formulas"].split(self.SPLIT_SIMBOL)
        student_formulas = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        coordinate = worksheet_str_ranges["student_formulas"].split(self.SPLIT_SIMBOL)
        point_formulas = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        coordinate = worksheet_str_ranges["average_formulas"].split(self.SPLIT_SIMBOL)
        average_formulas = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        average_formulas = tuple([row[0] for row in average_formulas])
        coordinate = worksheet_str_ranges["percentage_of_completion_formulas"].split(
            self.SPLIT_SIMBOL
        )
        percentage_of_completion_formulas = cell_finder.getCells(
            Range(coordinate[0], coordinate[-1])
        )
        percentage_of_completion_formulas = tuple(
            [row[0] for row in percentage_of_completion_formulas]
        )
        coordinate = worksheet_str_ranges["max_point_cells"].split(self.SPLIT_SIMBOL)
        max_point_cells = cell_finder.getCells(Range(coordinate[0], coordinate[-1]))
        max_point_cells = tuple([row[0] for row in max_point_cells])
        sum_max_point_formula: Cell = self.ws[
            worksheet_str_ranges["sum_max_point_formula"]
        ]
        coordinate = worksheet_str_ranges["sum_student_point_formulas"].split(
            self.SPLIT_SIMBOL
        )
        sum_student_point_formulas = cell_finder.getCells(
            Range(coordinate[0], coordinate[-1])
        )
        average_point: Cell = self.ws[worksheet_str_ranges["average_point"]]
        average_percentage_of_completion: Cell = self.ws[
            worksheet_str_ranges["average_percentage_of_completion"]
        ]
        coordinate = worksheet_str_ranges["percentage_of_points"].split(
            self.SPLIT_SIMBOL
        )
        percentage_of_points = cell_finder.getCells(
            Range(coordinate[0], coordinate[-1])
        )

        return WorksheetRanges(
            worksheet_str_ranges["name"],
            table_headers[0],
            task_formulas,
            student_formulas[0],
            point_formulas,
            average_formulas,
            percentage_of_completion_formulas,
            max_point_cells,
            sum_max_point_formula,
            sum_student_point_formulas[0],
            average_point,
            average_percentage_of_completion,
            percentage_of_points[0],
        )
